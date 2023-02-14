import sys
import time
import re
import os
import json
import csv
import xlrd
import openpyxl
from openpyxl.cell import MergedCell

from dtaf_core.iotg_plugin.iotg_base_testcase import IotgBaseTestCase
from dtaf_core.lib.dtaf_constants import Framework, OperatingSystems
from dtaf_core.providers.sut_os_provider import SutOsProvider
from dtaf_core.providers.ac_power import AcPowerControlProvider
from dtaf_core.providers.bios_menu import BiosSetupMenuProvider
from dtaf_core.providers.bios_menu import BiosBootMenuProvider
from dtaf_core.providers.provider_factory import ProviderFactory
from dtaf_core.providers.uefi_shell import UefiShellProvider
from dtaf_core.providers.bios_provider import BiosProvider

from src.lib.common_content_lib import CommonContentLib
from src.lib.content_base_test_case import ContentBaseTestCase
from src.lib.bios_util import BiosUtil
from src.lib import content_exceptions

class IotgTestPnp(IotgBaseTestCase):

    CMD_DOWNLOAD_XMLCLI_TOOL = r'wget -c -t 5 --no-check-certificate https://ubit-artifactory-ba.intel.com/artifactory/star_framework-ba-local/dtaf/xmlcli/6.6_xmlcli_windows_linux_Python2\&3.zip'

    CMD_NEX_PCIE_SHELL = r'export https_proxy=http://proxy-dmz.intel.com:912; git clone https://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@github.com/xiaoyang-sean/SPR-R-EEC-SWS-PnP_pcie.git; ' \
                         r'cd SPR-R-EEC-SWS-PnP_pcie; chmod 777 *xlsx'

    XLSX_FILE = r'SPR-EEC-PnP-NEX-pcie.xlsx'
    SWS_PnP_PATH = r'/root/SPR-R-EEC-SWS-PnP_pcie'
    NEX_REPORT_TEMPLATE = r'{}/{}'.format(SWS_PnP_PATH, XLSX_FILE)
    XLSX_REPORT_FILE = r'C:\Automation\dtaf_content\src\iotg_tests\pnp\{}'.format(XLSX_FILE)

    LOG_PCIE_DIR = r'C:\pnp_log\pcie'

    KPI_RESULT_CSV_FILE = r'C:\pnp_log\kpi_result.csv'

    KPI_LABEL = r'PCIe_PnP'
    TARGET_FLAG = 0

    REBOOT_TIMEOUT = 600

    def __init__(self, test_log, arguments, cfg_opts, bios_config_file_path=None):
        super(IotgTestPnp, self).__init__(test_log, arguments, cfg_opts)

        # Required for UEFI Shell testing
        # not required for Boot Menu and Setup Menu testing
        self.sut_os_cfg = cfg_opts.find(SutOsProvider.DEFAULT_CONFIG_PATH)
        self.os = ProviderFactory.create(self.sut_os_cfg, test_log)  # type: SutOsProvider

        self.arg_list = arguments

        bios_cfg = cfg_opts.find(BiosProvider.DEFAULT_CONFIG_PATH)
        self.bios = ProviderFactory.create(bios_cfg, test_log)  # type: BiosProvider

        # Required for bootmenu testing
        bootmenu_cfg = cfg_opts.find(BiosBootMenuProvider.DEFAULT_CONFIG_PATH) # type: BiosBootMenuProvider
        self.bootmenu = ProviderFactory.create(bootmenu_cfg, test_log)

        ac_power_cfg = cfg_opts.find(AcPowerControlProvider.DEFAULT_CONFIG_PATH)
        self.ac_power = ProviderFactory.create(ac_power_cfg, test_log)  # type: AcPowerControlProvider


        self._common_content_lib = CommonContentLib(self._log, self.os, cfg_opts)

        self.bios_config_file_path = bios_config_file_path
        self.bios_util = BiosUtil(cfg_opts,
                                  bios_config_file=self.bios_config_file_path,
                                  bios_obj=self.bios, common_content_lib=self._common_content_lib,
                                  log=self._log)

    @classmethod
    def add_arguments(cls, parser):
        super(IotgTestPnp, cls).add_arguments(parser)

        # sub module parameter: pcie
        parser.add_argument('--app', help="specific the sub module.")
        parser.add_argument('--qdf', help="specific the qdf.")
        parser.add_argument('--bit', help="specific the bit:512b/256b/128b/64b/32b.")
        parser.add_argument('--configfile', help="specific the pcie configfile.")
        parser.add_argument('--baseaddress', help="-a: specific an offset to all vector addresses.")
        parser.add_argument('--pcidomainnumber', help="-p: Specify the PCI Domain Number on which the Exerciser card is found (Linux only).")
        parser.add_argument('--time', help="-t: specific Time (in seconds) to run the test before automatically stopping.")
        parser.add_argument('--function', help="specific local/remote read/write: LR/LW/LRW/RR/RW/RRW.")
        parser.add_argument('--function1', help="specific read/write type: 0/1/2")


    def xlsx_get_target(self, case_name, baseadd, XLSX_SHEET_NAME):
        book = xlrd.open_workbook(self.XLSX_REPORT_FILE, 'rw')
        sheet = book.sheet_by_name(XLSX_SHEET_NAME)
        row_num = 0
        for i in range(sheet.nrows):
            if (case_name.lower() in sheet.row(i)[0].value.strip().replace('\n', '').lower()
                    and baseadd.lower() in sheet.row(i)[1].value.strip().replace('\n', '').lower()):
                row_num = i
                break
        #row_num = row_num + 1 if (self.arg_list.function1 == 1) else row_num
        #row_num = row_num if (self.arg_list.function1 == 2) else row_num + self.arg_list.function1
        target_col = 7
        target_value = {}
        # print(sheet.row(row_num)[target_col].value)
        self._log.info(f"Target : {sheet.row(row_num)[target_col].value}")
        if ("only" not in str(sheet.row(row_num)[target_col].value).lower()): # and ("Only" not in str(sheet.row(row_num)[target_col].value)):
            self.TARGET_FLAG = 1 # target
            target_value['Avg'] = round(float(sheet.row(row_num)[target_col].value), 2)
            target_value['Instant'] = 0
            target_value['Avg1'] = round(float(sheet.row(row_num + 1)[target_col].value), 2)
            target_value['Instant1'] = 0

        else:
            self.TARGET_FLAG = 0 # no target
            target_value['Avg'] = round(float(sheet.row(row_num)[6].value), 2)
            target_value['Instant'] = round(float(sheet.row(row_num)[5].value), 2)
            target_value['Avg1'] = round(float(sheet.row(row_num + 1)[6].value), 2)
            target_value['Instant1'] = round(float(sheet.row(row_num + 1)[5].value), 2)


        return target_value

    def create_log_dir(self, log_dir):
        is_exists = os.path.exists(log_dir)
        if not is_exists:
            os.makedirs(log_dir)

    def remove_sut_log_file(self, app, tag_file_name, log_script_file):
        self.os.execute(f"rm -rf {tag_file_name}", 100)
        self.os.execute(f"rm -rf {log_script_file}", 100)
        self.os.execute(r"rm -rf results", 100)
        self.os.execute(r"rm -rf emon", 100)

    def is_number(self, num):
        pattern = re.compile(r'^[-+]?[-0-9]\d*\.\d*|[-+]?\.?[0-9]\d*$')
        result = pattern.match(num)
        if result:
            return True
        else:
            return False

    def perform_os_reboot_with_tag(self, timeout):
        try:
            self.os.reboot(10)
        except Exception as ex:
            self._log.info("the exception '{}' is OK..".format(ex))

        start = time.time()
        self._log.info("Waiting for boot, up to {} seconds.".format(timeout))
        booted = False
        start_time = time.time()
        while not booted and time.time() - start_time < timeout:
            time.sleep(10)
            booted = self.os.is_alive()
            self._log.debug("SUT is " + ("alive!" if booted else "still not booted."))

        if not booted:
            self._log.info("SUT failed to boot within {} seconds!".format(timeout))
            return False

        end = time.time()
        total_time_taken = (abs(start - end))
        total_time_taken = ("{:05.2f}".format(total_time_taken))
        self._log.info("Enterd Into OS It Took {0} Seconds".format(total_time_taken))

        return True

    def graceful_reboot(self, timeout):
        if not self.perform_os_reboot_with_tag(timeout):
            #if soft reboot failed, then PDU power cycle
            self._log.info("Performs graceful shutdown SUT")
            self._common_content_lib.perform_graceful_ac_off_on(self.ac_power)
            self._common_content_lib.wait_for_os(timeout)

    def add_timestamp_to_file(self, log_script_file):
        timestamp = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime(time.time()))
        file_name = log_script_file.split('.')[0] + log_script_file.split('.')[1]
        ext_name = log_script_file.split('.')[2]

        file_name = file_name + '_' + timestamp

        return file_name + '.' + ext_name

    def check_system_date(self):
        # NUC's date
        localtime = time.localtime(time.time())

        # SUT's date
        year = self.os.execute(r"date +'%Y-%m-%d' | awk -F '-' '{{print$1}}'", 100)
        year = year.stdout.strip().replace('\n', '')
        self._log.info(f"year:{year}")

        month = self.os.execute(r"date +'%Y-%m-%d' | awk -F '-' '{{print$2}}'", 100)
        month = month.stdout.strip().replace('\n', '')
        self._log.info(f"month:{month}")

        day = self.os.execute(r"date +'%Y-%m-%d' | awk -F '-' '{{print$3}}'", 100)
        day = day.stdout.strip().replace('\n', '')
        self._log.info(f"day:{day}")

        if (not self.is_number(year)) or (not self.is_number(month)) or (not self.is_number(day)):
             self._log.error(f"year: {year} or month: {month} or day: {day} is not valid")
             return False

        self._log.info(f"Currently NUC date is:{localtime.tm_year}-{localtime.tm_mon}-{localtime.tm_mday}")
        if (int(localtime.tm_year) != int(year)) or (int(localtime.tm_mon) != int(month)) or (int(localtime.tm_mday) != int(day)):
            cur_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
            self._log.info(f"will set SUT date as: {cur_time}")
            self.os.execute(f"date -s '{cur_time}'", 100)
            self.os.execute(r"clock -w", 100)
            self._log.info("Set SUT date successfully")

        return True

    def print_memory_size(self):

        memory_size = self.os.execute("dmidecode | grep -A 5 'Memory Device' | grep Size | grep -v 'Range'", 1000)
        memory_size = memory_size.stdout if not memory_size.stderr else ''
        memory_size = memory_size.split('\n')

        total_memory_size = 0
        for size in memory_size:
            if size and ('No Module Installed' not in size):
                total_memory_size += int(size.split('Size: ')[-1].split('GB')[0].strip())

        self._log.info(f"Total Available Memory on SUT - {total_memory_size} GB")

        return True

    def append_result_to_csv(self, app, function, qdf, unit, data):
        timestamp = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime(time.time()))

        is_exists = os.path.exists(self.KPI_RESULT_CSV_FILE)
        if not is_exists:
            with open(self.KPI_RESULT_CSV_FILE, "a+") as csv_file:
                csv_file.writelines("TimeStamp,App,TestCase,QDF,Data['Instant'],Data['Avg'],Unit\n")
                csv_file.writelines("{},{},{},{},{},{},{}\n".format(timestamp, app, function, qdf, data['Instant'],data['Avg'], unit))
                csv_file.close()
                self._log.info("append CSV title and KPI result successfully")
        else:
            with open(self.KPI_RESULT_CSV_FILE, "a+") as csv_file:
                csv_file.writelines("{},{},{},{},{},{},{}\n".format(timestamp, app, function, qdf, data['Instant'],data['Avg'], unit))
                csv_file.close()
                self._log.info("append KPI result to csv successfully")

    def float_to_percentage(self, float_value, precision):
        value = float_value * 100
        return str(round(value, precision)) + '%'

    # get row value based on case_name
    #row_value = self.xlrd_get_row(self.XLSX_REPORT_FILE, casename, baseaddress, XLSX_SHEET_NAME)
    def xlrd_get_row(self, xlsx_file, case_name, baseaddress, sheet_name):
        book = xlrd.open_workbook(xlsx_file, 'rw')
        sheet = book.sheet_by_name(sheet_name)
        row_num = 0
        for i in range(sheet.nrows):
            if (case_name.lower() in sheet.row(i)[0].value.strip().replace('\n', '').lower()
                    and baseaddress.lower() in sheet.row(i)[1].value.strip().replace('\n', '').lower()):
                row_num = i
                break
        #row_num = row_num + 1 if (self.arg_list.function1 == 2) else row_num
        return row_num

    # swap two col value
    def openpyxl_swap_col_value(self, xlsx_file, col1, col2, sheet_name):
        workbook = openpyxl.load_workbook(xlsx_file, read_only=False)
        sheet = workbook[sheet_name]
        for i in range(sheet.max_row):
            cell_col1 = sheet.cell(row=i + 1, column=col1).value
            cell_col2 = sheet.cell(row=i + 1, column=col2).value
            sheet.cell(i + 1, col1, cell_col2)
            sheet.cell(i + 1, col2, cell_col1)
        workbook.save(xlsx_file)

    def openpyxl_write_cell(self, xlsx_file, row, col, cell_value, sheet_name):
        workbook = openpyxl.load_workbook(xlsx_file, read_only=False)
        sheet = workbook[sheet_name]
        sheet.cell(row, col, cell_value)
        workbook.save(xlsx_file)

    def openpyxl_set_col_default_value(self, xlsx_file, col, cell_value, sheet_name):
        workbook = openpyxl.load_workbook(xlsx_file, read_only=False)
        sheet = workbook[sheet_name]

        for i in range(sheet.max_row):
            cell = sheet.cell(row=i + 1, column=col)
            if isinstance(cell, MergedCell): # judge MergedCell
                continue
            if self.is_number(str(cell.value)):
                sheet.cell(i + 1, col, cell_value)
        workbook.save(xlsx_file)

    #self.print_kpi_results(app, function_str, socket_num, baseaddress, qdf, unit, data)
    def print_kpi_results(self, app, casename, socket_num, baseaddress, qdf, unit, data):
        self._log.info(f'print_results app: {app}, case name:{casename}, socket_num: {socket_num},'
                       f' baseaddress:{baseaddress}, qdf:{qdf}, unit: {unit}, data: {data}')

        try:
            gio_data = {}
            self.TARGET_FLAG = 0
            XLSX_SHEET_NAME = r'PCIE (CentOS) 1S' if socket_num == 1 else r'PCIE (CentOS) 2S'
            #excel_write_flag
            excel_flag = f'excel_write_flag'
            if self.os.check_if_path_exists(excel_flag):
                self.openpyxl_swap_col_value(self.XLSX_REPORT_FILE, 4, 6, XLSX_SHEET_NAME)
                self.openpyxl_swap_col_value(self.XLSX_REPORT_FILE, 5, 7, XLSX_SHEET_NAME)
                self._log.info(f'openpyxl_swap_col_value once !!!')
                self.os.execute(r'sudo rm -rf excel_write_flag', 30)
                self._log.info(f'execute cmd: sudo rm -rf excel_write_flag once !!!')
                self.openpyxl_set_col_default_value(self.XLSX_REPORT_FILE, 4, 'NA', XLSX_SHEET_NAME)
                self.openpyxl_set_col_default_value(self.XLSX_REPORT_FILE, 5, 'NA', XLSX_SHEET_NAME)
                self._log.info(f'openpyxl_set_col_default_value once !!!')

                # self.os.execute(r'sudo touch excel_write_flag', 30)
                # self._log.info("touch excel_write_flag successfully!")

            # get target value
            Target_value = self.xlsx_get_target(casename, baseaddress, XLSX_SHEET_NAME)
            self._log.info(f"Target_value : {Target_value}")
            gio_data['KPI_LABEL'] = ('_').join([self.KPI_LABEL, qdf, app]).lower()
            gio_data['app'] = app
            gio_data['TestCase'] = casename
            gio_data['Unit'] = unit
            gio_data['Data'] = data
            gio_data['Target_value'] = None if (self.TARGET_FLAG == 0) else Target_value

            tempdata = data['Avg1'] if (int(self.arg_list.function1) == 1) else data['Avg']
            temptarget = Target_value['Avg1'] if (int(self.arg_list.function1) == 1) else Target_value['Avg']

            gio_data['deviation percentage of target_value'] = None \
                if (self.TARGET_FLAG == 0) else self.float_to_percentage(float(tempdata / temptarget - 1), 2)
            gio_data['last_cycle_value'] = None if (self.TARGET_FLAG == 1) else Target_value
            gio_data['deviation percentage of last_cycle_value'] = None \
                if (self.TARGET_FLAG == 1) else self.float_to_percentage(float(tempdata / temptarget - 1), 2)
            print(gio_data['deviation percentage of last_cycle_value'])

            #gio_data['Target_value'] = Target_value
            Target_pass_condition = -0.01 if (self.TARGET_FLAG == 1) else -0.05
            pass_condition = float(data['Avg1'] / Target_value['Avg1'] - 1) if (int(self.arg_list.function1) == 1) else float(data['Avg'] / Target_value['Avg'] - 1)
            gio_data['Pass_or_fail'] = 'pass' if (pass_condition > Target_pass_condition) else f'encounter ~{Target_pass_condition* 100}% performance drop on the BKC !'
            # gio_data['Pass_or_fail'] = 'pass' \
            #     if (-0.35 < float(data / Target_value - 1) < 0.35) else 'encounter ~35% performance drop on the BKC !'
            one_line_res=[gio_data['KPI_LABEL'], gio_data['TestCase'], gio_data['Unit'], gio_data['Data'],
                          gio_data['Target_value'], gio_data['deviation percentage of target_value'],
                          gio_data['last_cycle_value'], gio_data['deviation percentage of last_cycle_value'], gio_data['Pass_or_fail']]
            print(one_line_res)

            # write data to excel file
            row_value = self.xlrd_get_row(self.XLSX_REPORT_FILE, casename, baseaddress, XLSX_SHEET_NAME)
            self.openpyxl_write_cell(self.XLSX_REPORT_FILE, row_value + 1, 5, data['Avg'], XLSX_SHEET_NAME)
            self.openpyxl_write_cell(self.XLSX_REPORT_FILE, row_value + 1, 4, data['Instant'], XLSX_SHEET_NAME)
            self.openpyxl_write_cell(self.XLSX_REPORT_FILE, row_value + 2, 5, data['Avg1'], XLSX_SHEET_NAME)
            self.openpyxl_write_cell(self.XLSX_REPORT_FILE, row_value + 2, 4, data['Instant1'], XLSX_SHEET_NAME)


        except:
            print('**NOT SUPPORT YET**')
            gio_data = {}
            one_line_res = []
            pass

        print('KPI_JSON_START')
        print(json.dumps(gio_data, indent=4))
        print('KPI_JSON_END')
        print()

        # append the result to CSV file
        self.append_result_to_csv(app, casename, qdf, unit, data)
        if gio_data['Pass_or_fail'] != 'pass' :
            raise Exception(gio_data['Pass_or_fail'])

        return gio_data

    def get_socket_number(self):
        checkout = self.os.execute(r"lscpu | awk '/Socket\(s\):/{{print $NF}}'", 1000)
        return int(checkout.stdout.strip().replace('\n', ''))

    def parse_pcie_kpi_data(self, app, bit, function, socket_num, baseaddress, qdf, log_file):
        grep_str = 'Reads'
        grep_str1 = 'Writes'
        data = {}
        unit = 'MB/s'
        func = ''

        # parser.add_argument('--function', help="specific local/remote read/write: LR/LW/LRW/RR/RW/RRW.")
        if (function == 'LR') or (function == 'RR'):
            func = 'Read'
            #grep_str = 'Reads'
        if (function == 'LW') or (function == 'RW'):
            func = 'Write'
            #grep_str = 'Writes'
        if (function == 'LRW') or (function == 'RRW'):
            func = 'Read & Write'
            #grep_str = 'Reads' #if (self.arg_list.function1 == 1) else 'Writes'

        pcie_Gen = "Gen4" if (socket_num == 1) else "Gen5"
        slot_name = 'Slot_E' if (int(self.arg_list.pcidomainnumber) == 152) else 'Slot_B'
        if (function == 'LR') or (function == 'LW') or (function == 'LRW'):
            function_str = r'IO 1x16 PCIe {} BW {} {} ({})'.format(pcie_Gen, bit, func, slot_name)
        else:
            function_str = r'Remote IO 1x16 PCIe {} BW {} {} ({})'.format(pcie_Gen, bit, func, slot_name)

        # get Avg value
        cmd_Avg = r"cat {} | grep '{}' | tail -n 1 | awk -F 'MB/s' '{{print $2}}'".format(log_file, grep_str)
        self._log.info(f"cmd:{cmd_Avg}")
        checkout = self.os.execute(cmd_Avg, 100)
        self._log.info(f"data['Avg'] value:{checkout.stdout}")
        data['Avg'] = checkout.stdout.strip().replace('\n', '')

        # get instant value
        cmd_Instant = r"cat {} | grep '{}' | tail -n 1 | awk -F 'MB/s' '{{print $1}}' | awk -F ':' '{{print $2}}'".format(log_file, grep_str)
        self._log.info(f"cmd_instant:{cmd_Instant}")
        checkout = self.os.execute(cmd_Instant, 100)
        self._log.info(f"data['instant'] value:{checkout.stdout}")
        data['Instant'] = checkout.stdout.strip().replace('\n', '')


        cmd_Avg1 = r"cat {} | grep '{}' | tail -n 1 | awk -F 'MB/s' '{{print $2}}'".format(log_file, grep_str1)
        self._log.info(f"cmd:{cmd_Avg1}")
        checkout = self.os.execute(cmd_Avg1, 100)
        self._log.info(f"Avg1 value:{checkout.stdout}")
        data['Avg1'] = checkout.stdout.strip().replace('\n', '')

        # get instant value
        cmd_Instant1 = r"cat {} | grep '{}' | tail -n 1 | awk -F 'MB/s' '{{print $1}}' | awk -F ':' '{{print $2}}'".format(log_file, grep_str1)
        self._log.info(f"cmd_instant:{cmd_Instant1}")
        checkout = self.os.execute(cmd_Instant1, 100)
        self._log.info(f"data['instant1'] value:{checkout.stdout}")
        data['Instant1'] = checkout.stdout.strip().replace('\n', '')


        # check the data validation, NULL or 0.00 will be failed
        # if data['Avg'] == '' or data['Instant'] == '' or (not self.is_number(data['Instant'])) \
        #         or (not self.is_number(data['Avg'])) or data['Avg1'] == '' or data['Instant1'] == '':
        if data['Avg'] == '' or data['Instant'] == '' \
                 or data['Avg1'] == '' or data['Instant1'] == '':
            self._log.error(f"Failed to get the pcie data, case name: {function_str},"
                            f" data['Instant']: {data['Instant']}, data['Avg']: {data['Avg']},"
                            f" data['Instant1']: {data['Instant1']}, data['Avg1']: {data['Avg1']}")
            # even failed, still need output the JSON KPI data
            self.print_kpi_results(app, function_str, socket_num, baseaddress, qdf, unit, data)

            return False


        data['Instant'] = round(float(data['Instant']), 2)
        data['Avg'] = round(float(data['Avg']), 2)
        data['Instant1'] = round(float(data['Instant1']), 2)
        data['Avg1'] = round(float(data['Avg1']), 2)

        self.print_kpi_results(app, function_str, socket_num, baseaddress, qdf, unit, data)

        return True

    def bios_setting(self):
        perform_os_reboot_flag = False
        #VT-D and Extended APIC are bundled, please disable Extended APIC first when you want to disable VTD
        #Socket Configuration -> IIO Configuration -> Intel VT for Directed I/O (VT-d) -> Intel VT for Directed I/O
        #Socket Configuration -> Processor Configuration -> Extended APIC
        Extended_APIC = self.bios_util.get_bios_knob_current_value("ProcessorX2apic")
        self._log.info("Extended_APIC value={}".format(Extended_APIC))

        if int(Extended_APIC, 16) != int("0x0", 16):
            self.bios_util.set_single_bios_knob("ProcessorX2apic", "0x0")
            self._log.info("set/disable Extended_APIC value: 0x0")
            perform_os_reboot_flag =  True

        VTdSupport = self.bios_util.get_bios_knob_current_value("VTdSupport")
        self._log.info("VTdSupport value={}".format(VTdSupport))

        if int(VTdSupport, 16) != int("0x0", 16):
            self.bios_util.set_single_bios_knob("VTdSupport", "0x0")
            self._log.info("set/disable VTdSupport value: 0x0")
            perform_os_reboot_flag =  True

        return perform_os_reboot_flag

    def prepare(self):
        if self.os.os_type != OperatingSystems.LINUX:
            self._log.error("This test case only applicable for linux systems")
            return False

        # pre-setup the SUT
        # 1. install python3/git tool if necessary, for RedHat, it is not required, if CentOS, then required.
        # check the certifaction file exist, if not, then copy
        if not self.os.check_if_path_exists("/etc/pki/tls/certs/ca-bundle.crt", False):
            #crt_path = self.os.execute("cd /etc; find -name 'ca-bundle.crt' | tail -n 1", 1000).stdout.strip().replace('\n', '')
            #if 'ca-bundle.crt' in crt_path:
                #self.os.execute(f"cd /etc; sudo copy {crt_path} ./pki/tls/certs/ca-bundle.crt", 1000)
            self.os.execute(r"sudo copy /etc/pki/ca-trust/extracted/pem/tls-ca-bundle.pem /etc/pki/tls/certs/ca-bundle.crt", 1000)

        # check the system date, if too old, then need set the date, otherwise will happen SSL certificate problem: certificate is not yet valid
        if not self.check_system_date():
            self._log.error("Set System Date failed")
            return False

        # 2. download xmlcli tool from SUT 
        if not self.os.check_if_path_exists("/opt/APP/xmlcli", True):
            result = self._common_content_lib.execute_sut_cmd(sut_cmd=self.CMD_DOWNLOAD_XMLCLI_TOOL,cmd_str="Download the XMLCLI tool",execute_timeout=1000)
            if result != "":
                err_msg = "Fail to Prepare xmlcli tool on SUT"
                output = "{}:{}".format(err_msg, result)
                self._log.error(output)
                return False

            if not self.os.check_if_path_exists("/opt/APP", True):
                self.os.execute("sudo mkdir /opt/APP/", 1000)

            self.os.execute("sudo mv *xmlcli*.zip /opt/APP/", 1000)
            self.os.execute("cd /opt/APP/; unzip *xmlcli*.zip; mv *xmlcli*/xmlcli ./", 1800)
            self._log.info("Download the xmlcli tool and unzip successfully")

        if self.os.check_if_path_exists(self.SWS_PnP_PATH, True):
            self._log.info("{} exist, not need to clone report file!")
            return True

        check = self.os.execute(f"{self.CMD_NEX_PCIE_SHELL}", 1200)
        self._log.info(check.stdout)

        if not self.os.check_if_path_exists(f"{self.NEX_REPORT_TEMPLATE}", False):
            self._log.error("{} not exist!".format(self.NEX_REPORT_TEMPLATE))
            return False

        self.os.copy_file_from_sut_to_local(self.NEX_REPORT_TEMPLATE, self.XLSX_REPORT_FILE)
        self._log.info('succeed to copy report file from sut to NUC.')

        self.os.execute(r'sudo touch excel_write_flag', 30)
        self._log.info("touch excel_write_flag successfully!")
        return True

    def execute(self):
        try:
            self._log.info("******PnP Task Start******")

            #execute_status = False
            parse_status = True

            log_script_file = f'{self.arg_list.app}_{self.arg_list.configfile}_{self.arg_list.qdf}_{self.arg_list.function}_{self.arg_list.baseaddress}.log'
            #tag_file_name = f'{self.arg_list.app}_{self.arg_list.configfile}_{self.arg_list.qdf}_{self.arg_list.function}_{self.arg_list.baseaddress}.flag'
            log_host_dir = self.LOG_PCIE_DIR


            # print memory infomation
            self.print_memory_size()

            # BIOS knobs setting if need
            perform_os_reboot_flag = self.bios_setting()
            if perform_os_reboot_flag:
                self.graceful_reboot(self.REBOOT_TIMEOUT)

            self._log.info(f"will execute {self.arg_list.app} module.")
            self._log.info(f"configfile: {self.arg_list.configfile}, function: {self.arg_list.function}, function1:{self.arg_list.function1}.")
            socket_num = self.get_socket_number()
            self.create_log_dir(log_host_dir)
            # if (self.arg_list.function1 == 1):
            #     self.os.execute(f"sudo touch {tag_file_name}", 100)
            #
            # if (self.arg_list.function1 == 2) and self.os.check_if_path_exists(tag_file_name) and self.os.check_if_path_exists(log_script_file):
            #     if not self.parse_pcie_kpi_data(self.arg_list.app, self.arg_list.bit, self.arg_list.function,
            #                                     socket_num,
            #                                     self.arg_list.baseaddress, self.arg_list.qdf, log_script_file):
            #         self._log.error("Parse KPI data failed")
            #         return False
            #     self._log.info(f"configfile: {self.arg_list.configfile}, function: {self.arg_list.function} have executed.")
            #
            #     self.os.copy_file_from_sut_to_local(log_script_file, nucpath)
            #     self.os.execute(f"rm -rf {log_script_file}", 100)
            #     self.os.execute(f"rm -rf {tag_file_name}", 100)
            #
            #     return True

            # ececute pcie test command
            check = self.os.execute(
                f"cd /root/PTG && ./ptg 1 gens/sequential/{self.arg_list.configfile} -E -a {self.arg_list.baseaddress}"
                f" -p {self.arg_list.pcidomainnumber} -A -t 10 > /root/{log_script_file}",
                3600)
            self._log.info(check.stdout)

            # will directly KPI parser
            if not self.parse_pcie_kpi_data(self.arg_list.app, self.arg_list.bit, self.arg_list.function, socket_num,
                                            self.arg_list.baseaddress, self.arg_list.qdf, log_script_file):
                self._log.error("Parse KPI data failed")
                parse_status = False

            host_log_file = self.add_timestamp_to_file(log_script_file)
            nucpath = self.LOG_PCIE_DIR + f'\\{host_log_file}'
            self.os.copy_file_from_sut_to_local(log_script_file, nucpath)
            self.os.execute(f"rm -rf {log_script_file}", 100)

        except Exception as ex:
            self._log.error(ex)
            return False

        return parse_status


if __name__ == "__main__":
    sys.exit(Framework.TEST_RESULT_PASS if IotgTestPnp.main() else Framework.TEST_RESULT_FAIL)
